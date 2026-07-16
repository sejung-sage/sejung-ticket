#!/usr/bin/env python3
"""
반포 주간 시간표 XLSX → JSON (요일 시트 7개).
사용: python3 scripts/parse_timetable_xlsx.py "<xlsx경로>" [연도=2026]
출력(stdout): [{ "date":"2026-01-19", "weekday":1, "cells":[{room_raw,teacher,time_raw,detail},...] }, ...]

시트 구조: r1 날짜("1월 19일 월요일"), r2 관(병합: 본관/2관/SKY관), r3 호실("301\n(정원)"),
r5부터 30분/행 그리드(r5=9:00), 수업은 세로 병합 셀. 셀 텍스트에 "9:00-\n12:30"식
명시 시간이 있으면 우선하고(AM/PM은 그리드 행 위치로 해소), 없으면 병합 범위로 계산.
time_raw는 24시간제 "HH:MM-HH:MM"로 내보내 DB parse_time_range가 모호성 없이 파싱한다.
의존: openpyxl
"""
import sys, re, json, datetime
import openpyxl

DOW = {"월": 1, "화": 2, "수": 3, "목": 4, "금": 5, "토": 6, "일": 7}
GRID_START_ROW = 5          # r5 = 9:00
GRID_START_MIN = 9 * 60
SLOT_MIN = 30
TEACHER_RE = re.compile(r"([가-힣]{2,4})\s*T")
DATE_RE = re.compile(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일")
# "9:00- 12:30" / "8-9" 등. 끝 숫자 바로 뒤 '/'는 날짜 표기("19/26")이므로 제외하되,
# "9:30 1/19 종강"처럼 뒤에 딴 숫자가 와도 ':30'이 백트래킹으로 잘리지 않게 [:\d] 직후만 금지.
TIME_RE = re.compile(r"(\d{1,2})(?::(\d{2}))?\s*[~\-.]+\s*(\d{1,2})(?::(\d{2}))?(?!\s*/)(?![:\d])")


def resolve_hour(h, mn, grid_min):
    """12시간제 시각을 그리드 위치에 가장 가까운 24시간제로 해소."""
    cands = [h * 60 + mn]
    if h < 12:
        cands.append((h + 12) * 60 + mn)
    return min(cands, key=lambda m: abs(m - grid_min))


def cell_time(text, grid_start, grid_end):
    """명시 시간 우선(그리드로 AM/PM 해소), 없으면 그리드 범위."""
    m = TIME_RE.search(text)
    if m:
        h1, mn1 = int(m.group(1)), int(m.group(2) or 0)
        h2, mn2 = int(m.group(3)), int(m.group(4) or 0)
        start = resolve_hour(h1, mn1, grid_start)
        end = resolve_hour(h2, mn2, grid_end)
        if end <= start:
            end += 12 * 60
        if end <= start or end > 24 * 60:
            start, end = grid_start, grid_end
    else:
        start, end = grid_start, grid_end
    return f"{start // 60}:{start % 60:02d}-{end // 60}:{end % 60:02d}"


def parse_sheet(ws, sheet_name, year):
    m = DATE_RE.search(str(ws.cell(1, 1).value or ""))
    if not m:
        raise ValueError(f"[{sheet_name}] r1에서 날짜를 못 찾음: {ws.cell(1,1).value!r}")
    date = datetime.date(year, int(m.group(1)), int(m.group(2)))
    weekday = DOW[sheet_name.strip()]
    if date.isoweekday() != weekday:
        raise ValueError(f"[{sheet_name}] {date}는 isodow {date.isoweekday()} — 시트 요일과 불일치 (연도 확인)")

    # r2 관 헤더(병합) → 열별 관 이름 전개
    building_by_col = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= 2 <= mr.max_row:
            v = ws.cell(2, mr.min_col).value
            if isinstance(v, str) and v.strip():
                name = re.sub(r"\s+", "", v)
                for c in range(mr.min_col, mr.max_col + 1):
                    building_by_col[c] = name
    # r3 호실 헤더 → room_raw ("본관 301")
    room_by_col = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(3, c).value
        rm = re.match(r"\s*(\d+)", str(v or ""))
        if rm and c in building_by_col:
            room_by_col[c] = f"{building_by_col[c]} {rm.group(1)}"

    # 그리드 범위: A열 시간축 마지막 라벨 행 + 1
    axis_rows = [r for r in range(GRID_START_ROW, ws.max_row + 1) if ws.cell(r, 1).value is not None]
    grid_last_row = (max(axis_rows) if axis_rows else GRID_START_ROW) + 1

    span = {}  # (r,c) → merge max_row
    for mr in ws.merged_cells.ranges:
        span[(mr.min_row, mr.min_col)] = mr.max_row

    cells = []
    for r in range(GRID_START_ROW, grid_last_row + 1):
        for c in sorted(room_by_col):
            v = ws.cell(r, c).value
            if not isinstance(v, str) or not v.strip():
                continue
            text = re.sub(r"\s+", " ", v).strip()
            t = TEACHER_RE.findall(text)
            if not t:
                continue  # 안내문/전화번호 등
            end_row = min(span.get((r, c), r), grid_last_row)
            grid_start = GRID_START_MIN + (r - GRID_START_ROW) * SLOT_MIN
            grid_end = GRID_START_MIN + (end_row - GRID_START_ROW + 1) * SLOT_MIN
            cells.append({
                "room_raw": room_by_col[c],
                "teacher": t[0],
                "time_raw": cell_time(text, grid_start, grid_end),
                "detail": text[:120],
            })
    return {"date": date.isoformat(), "weekday": weekday, "cells": cells}


def main():
    path = sys.argv[1]
    year = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
    wb = openpyxl.load_workbook(path, data_only=True)
    days = [parse_sheet(wb[n], n, year) for n in wb.sheetnames if n.strip() in DOW]
    print(json.dumps(days, ensure_ascii=False))


if __name__ == "__main__":
    main()
